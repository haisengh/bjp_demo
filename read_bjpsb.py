from openpyxl import Workbook, load_workbook
from pymongo import MongoClient
ws = load_workbook("test.xlsx")["sheet"]

def extarct(row):
	collection = []
	for cell in row:
		if cell.value:
			collection.append({"name": cell.value, "span": [cell,]})
		else:
			collection[-1]["span"].append(cell)
	return collection

def span(span):
	"""
	return the tuple following the cells specified by the span.
	"""
	start = span[0].col_idx - 1
	end = span[-1].col_idx
	row = span[0].row + 1
	return ws[row][start:end]




def table(cate):
	"""
	return the table with complete elements spanned by "span".
	cate is the dict formed by the table name.
	"""
	def create_table(th):
		"""
		th ::= table head. a tuple consist of two cells.
		return the collection.
		"""
		tables = []
		row,col = th[0].row, th[0].column
		row1,col1 = th[1].row, th[1].column
		li = ws[col][row:]
		li1 = ws[col1][row1:]
		for a,b in zip(li,li1):
			if a.value == None:
				return tables
			tables.append({"name": a.value, "time": b.value, "row": a.row})
			
	properties = extarct(span(cate["span"])) #verbs, nouns, special_nouns,
	ul = [span(properti["span"]) for properti in properties]
	return {"name": cate["name"], "verb_tables": create_table(ul[0]), "noun_tables": create_table(ul[1]), "special_noun": create_table(ul[2])}
#字典似乎不是以引用的方式传值的。

def main():
	client = MongoClient("mongodb://127.0.0.1:27017/")	
	cates = extarct(ws[1])
	blockwords = []
	for n,i in enumerate(cates):
		print(i["name"],n)
		blockwords.append(table(i))
	db = client["BJPSB"]
	collection = db["blockwords"]
	result = collection.insert_many(blockwords)
	print(result)

if __name__ == '__main__':
	main()






