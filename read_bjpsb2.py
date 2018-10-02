from openpyxl import Workbook, load_workbook
ws = load_workbook("test.xlsx")["sheet"]

def extarct(row):
	collection = []
	for cell in row:
		if cell.value:
			collection.append({"name": cell.value, "span": [cell,]})
		else:
			collection[-1]["span"].append(cell)
	return collection

def row(cate):
	span = cate["span"]
	col_left = span[0].column
	init = span[0].col_idx - 1
	end = span[-1].col_idx - 1
	for i in ws[col_left]:
		if i.value != None:
			row = i.row
			yield ws[row][init:end-1]

cates = extarct(ws[1])
cate = cates[0]

for i in row(cate):
	print(i)